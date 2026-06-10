"""
Build an enriched VCDP Bulk Upload Excel file that:
1. Reads data from the existing populated Excel
2. Maps fields to the template column structure
3. Applies all dropdown validations using openpyxl
4. Outputs a ready-to-upload file

Run from: backend/
  uv run python build_populated_template.py
"""
import asyncio
import io
import json
import os
import sys
import pandas as pd
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import selectinload

sys.path.append(os.path.abspath('.'))

SOURCE_FILE = r"../excel/Test_Bulk_Upload_Populated.xlsx"
OUTPUT_FILE = r"../excel/VCDP_Populated_Template.xlsx"

TEMPLATE_COLUMNS = [
    "Ref ID", "Project / Activity Name", "Activity Name", "Activity Code", "Category / Costcode",
    "Record Type", "Status", "Institution Code", "Executing Agency", "FY Awarded",
    "FY Completed", "Programme Phase", "Fiscal Quarter", "LGA(s)", "Commodity",
    "VCDP Component", "VCDP Sub-Component(s)", "3FS Primary Component", "3FS Sub-Component(s)",
    "COFOG Code", "COFOG Division(s)", "COFOG Group(s)", "Funding Sources", "Sub Funding Sources",
    "Currency", "Exchange Rate", "Expenditure – FGN", "Expenditure - State",
    "Expenditure - IFAD Loan", "Expenditure - IFAD Grant", "Expenditure - OOF",
    "Expenditure - Beneficiary", "Expenditure - Private Sector",
    "Expenditure - Value Chain", "Expenditure - Other", "Expenditure - Total Reported",
    "Beneficiary Unit", "Beneficiaries – Total", "Beneficiaries - Male", "Beneficiaries - Female",
    "Beneficiaries - Youth (<35)", "Beneficiaries - PLWD", "Beneficiary Categories",
    "Quantity Q1", "Quantity Q2", "Quantity Q3", "Quantity Q4",
    "Value Chain Segment(s)", "Value Chain Segments (Other)", "Climate Aligned?",
    "Data Source(s)", "Classification Notes"
]

VALID_STATE_TABS = [
    'Anambra', 'Benue', 'Ebonyi', 'Enugu', 'Kogi',
    'Nasarawa', 'Niger', 'Ogun', 'Taraba', 'NPMU'
]

VCDP_COMPONENTS = {
    "Component 1: Agricultural Market Development": [
        "Value Addition & Market Linkages", "Market Infrastructure",
        "Mainstreaming activities such as gender/youth, financial inclusion, E&CC and nutrition are cross-cutting",
    ],
    "Component 2: Smallholder Productivity Enhancement": [
        "Strengthening Farmers' Organisations", "Smallholder Production",
        "Mainstreaming activities such as gender/youth, financial inclusion, E&CC and nutrition are cross-cutting"
    ],
    "Component 3: Programme Management and Coordination": [
        "Procurement", "Knowledge Management & Sharing", "M&E", "Finance/Audit", "Admin",
    ],
}

THREEFS_COMPONENTS = {
    "1. Food Production": ["Production support (On-farm)", "Input supply & technologies", "Extension services"],
    "2. Food processing": ["Agro-processing", "Processing centers"],
    "3. Food social protection": ["Cash transfer schemes", "Emergency food assistance", "School feeding programs"],
    "4. Enabling environment": ["Rural roads or other transportation networks", "Market linkages"],
    "5. Governance": ["Climate adaptation", "Land management"],
}

FUNDING_SOURCES = {
    "Domestic Public Financing": ["FGN counterpart funding", "State/LGA contribution", "Beneficiary Contribution"],
    "International Development Financing": [
        "IFAD loan (Official Development Assistance (ODA))",
        "IFAD Grants (Official Development Assistance (ODA))", "Other Official Flows (OOF)"
    ],
    "Private Sector Financing": [],
    "Value Chain Financing": [],
}

VALUE_CHAIN_SEGMENTS = ["Production", "Processing", "Marketing", "Others"]
COMMODITIES = ["Rice", "Cassava", "Cross-cutting"]
FISCAL_YEARS = list(range(2013, 2051))


def fmt_list(val):
    """Try to parse a list field, return as comma-separated string."""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.startswith('['):
        try:
            items = json.loads(s)
            return ", ".join(str(i) for i in items)
        except:
            pass
    return s


async def build():
    from app.database import AsyncSessionLocal
    from app.models import State

    # Load source data
    src = pd.ExcelFile(SOURCE_FILE)
    print(f"Source sheets: {src.sheet_names}")

    # Map source column names to template names
    COL_MAP = {
        "Ref ID": "Ref ID",
        "Project / Activity Name": "Project / Activity Name",
        "Activity Name": "Activity Name",
        "Activity Code": "Activity Code",
        "Category / Costcode": "Category / Costcode",
        "Record Type": "Record Type",
        "Status": "Status",
        "Institution Code": "Institution Code",
        "Executing Agency": "Executing Agency",
        "FY Awarded": "FY Awarded",
        "FY Completed": "FY Completed",
        "Programme Phase": "Programme Phase",
        "Fiscal Quarter": "Fiscal Quarter",
        "LGA(s)": "LGA(s)",
        "Commodity": "Commodity",
        "VCDP Component": "VCDP Component",
        "VCDP Sub-Component(s)": "VCDP Sub-Component(s)",
        "3FS Primary Component": "3FS Primary Component",
        "3FS Sub-Component(s)": "3FS Sub-Component(s)",
        "COFOG Code": "COFOG Code",
        "COFOG Division(s)": "COFOG Division(s)",
        "COFOG Group(s)": "COFOG Group(s)",
        "Funding Sources": "Funding Sources",
        "Sub Funding Sources": "Sub Funding Sources",
        "Currency": "Currency",
        "Exchange Rate": "Exchange Rate",
        "Expenditure – FGN": "Expenditure – FGN",
        "Expenditure - State": "Expenditure - State",
        "Expenditure - IFAD Loan": "Expenditure - IFAD Loan",
        "Expenditure - IFAD Grant": "Expenditure - IFAD Grant",
        "Expenditure - OOF": "Expenditure - OOF",
        "Expenditure - Beneficiary": "Expenditure - Beneficiary",
        "Expenditure - Private Sector": "Expenditure - Private Sector",
        "Expenditure - Value Chain": "Expenditure - Value Chain",
        "Expenditure - Other": "Expenditure - Other",
        "Expenditure - Total Reported": "Expenditure - Total Reported",
        "Beneficiary Unit": "Beneficiary Unit",
        "Beneficiaries – Total": "Beneficiaries – Total",
        "Beneficiaries - Male": "Beneficiaries - Male",
        "Beneficiaries - Female": "Beneficiaries - Female",
        "Beneficiaries - Youth (<35)": "Beneficiaries - Youth (<35)",
        "Beneficiaries - PLWD": "Beneficiaries - PLWD",
        "Beneficiary Categories": "Beneficiary Categories",
        "Quantity Q1": "Quantity Q1",
        "Quantity Q2": "Quantity Q2",
        "Quantity Q3": "Quantity Q3",
        "Quantity Q4": "Quantity Q4",
        "Value Chain Segment(s)": "Value Chain Segment(s)",
        "Value Chain Segments (Other)": "Value Chain Segments (Other)",
        "Climate Aligned?": "Climate Aligned?",
        "Data Source(s)": "Data Source(s)",
        "Classification Notes": "Classification Notes",
    }

    # Load all source data grouped by sheet
    data_by_state = {}
    for sheet in src.sheet_names:
        mapped_sheet = "NPMU" if sheet in ("FCT", "FCT (NPMU)", "NPMU") else sheet
        if mapped_sheet not in VALID_STATE_TABS:
            continue
        df = pd.read_excel(src, sheet_name=sheet)
        if df.empty:
            continue
        # Remap columns
        rows = []
        for _, row in df.iterrows():
            new_row = {col: "" for col in TEMPLATE_COLUMNS}
            for src_col, tgt_col in COL_MAP.items():
                if src_col in df.columns:
                    val = row.get(src_col, "")
                    # Format list fields nicely
                    if tgt_col in ["Commodity", "VCDP Component", "3FS Primary Component",
                                   "Fiscal Quarter", "LGA(s)", "Funding Sources",
                                   "Value Chain Segment(s)", "Beneficiary Categories",
                                   "VCDP Sub-Component(s)", "3FS Sub-Component(s)",
                                   "COFOG Division(s)", "COFOG Group(s)"]:
                        new_row[tgt_col] = fmt_list(val)
                    else:
                        new_row[tgt_col] = val if not pd.isna(val) else ""
            rows.append(new_row)
        data_by_state[mapped_sheet] = rows
        print(f"  Loaded {len(rows)} rows from sheet '{sheet}' -> '{mapped_sheet}'")

    # Load LGAs from DB
    async with AsyncSessionLocal() as db:
        result = await db.execute(select(State).options(selectinload(State.lgas)))
        db_states = result.scalars().all()
        state_lga_map = {s.name: [l.name for l in s.lgas] for s in db_states}
        if "FCT (NPMU)" in state_lga_map:
            state_lga_map["NPMU"] = state_lga_map["FCT (NPMU)"]

    # Build the Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        options_ws = workbook.create_sheet("Options")
        options_ws.sheet_state = 'hidden'

        options_map = {
            "RecordType": ["Actual", "Budget"],
            "Status": ["DRAFT", "PENDING", "PUBLISHED", "REJECTED"],
            "Currency": ["NGN", "USD"],
            "Climate": ["Yes", "No"],
            "Commodity": COMMODITIES,
            "Quarters": ["Q1", "Q2", "Q3", "Q4", "Q1, Q2, Q3, Q4"],
            "VCDP_Comp": list(VCDP_COMPONENTS.keys()),
            "3FS_Comp": list(THREEFS_COMPONENTS.keys()),
            "Years": [str(y) for y in FISCAL_YEARS],
            "Units": ["Person", "Hectares", "Kilometers", "Tons", "Number", "Other"],
            "Phases": ["Original (2013-2018)", "1st AF", "2nd AF"],
            "Funding": list(FUNDING_SOURCES.keys()),
            "Segments": VALUE_CHAIN_SEGMENTS,
        }

        col_idx = 1
        ref_map = {}
        for key, vals in options_map.items():
            for row_idx, val in enumerate(vals, 1):
                options_ws.cell(row=row_idx, column=col_idx, value=val)
            ref_map[key] = f"'Options'!${get_column_letter(col_idx)}$1:${get_column_letter(col_idx)}${len(vals)}"
            col_idx += 1

        lga_ref_map = {}
        for state_name, lgas in state_lga_map.items():
            if not lgas:
                continue
            for row_idx, lga in enumerate(lgas, 1):
                options_ws.cell(row=row_idx, column=col_idx, value=lga)
            lga_ref_map[state_name] = f"'Options'!${get_column_letter(col_idx)}$1:${get_column_letter(col_idx)}${len(lgas)}"
            col_idx += 1

        # Write each state tab
        for state in VALID_STATE_TABS:
            rows = data_by_state.get(state, [])
            if rows:
                df_out = pd.DataFrame(rows, columns=TEMPLATE_COLUMNS)
            else:
                df_out = pd.DataFrame(columns=TEMPLATE_COLUMNS)

            df_out.to_excel(writer, index=False, sheet_name=state)
            worksheet = writer.sheets[state]

            def add_val(col_name, formula):
                if not formula:
                    return
                try:
                    idx = TEMPLATE_COLUMNS.index(col_name) + 1
                    col_letter = get_column_letter(idx)
                    dv = DataValidation(
                        type="list",
                        formula1=formula,
                        allow_blank=True,
                        showErrorMessage=True,
                        errorTitle="Invalid Selection",
                        error="Please select a value from the dropdown list."
                    )
                    dv.add(f"{col_letter}2:{col_letter}2000")
                    worksheet.add_data_validation(dv)
                except ValueError:
                    pass

            add_val("Record Type", ref_map["RecordType"])
            add_val("Status", ref_map["Status"])
            add_val("Currency", ref_map["Currency"])
            add_val("Climate Aligned?", ref_map["Climate"])
            add_val("Commodity", ref_map["Commodity"])
            add_val("Fiscal Quarter", ref_map["Quarters"])
            add_val("VCDP Component", ref_map["VCDP_Comp"])
            add_val("3FS Primary Component", ref_map["3FS_Comp"])
            add_val("FY Awarded", ref_map["Years"])
            add_val("FY Completed", ref_map["Years"])
            add_val("Beneficiary Unit", ref_map["Units"])
            add_val("Programme Phase", ref_map["Phases"])
            add_val("Funding Sources", ref_map["Funding"])
            add_val("Value Chain Segment(s)", ref_map["Segments"])
            add_val("LGA(s)", lga_ref_map.get(state, ""))

            # Style: freeze header row
            worksheet.freeze_panes = "A2"

            # Auto-adjust column widths
            for col in worksheet.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value),
                    default=8
                )
                worksheet.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

            print(f"  OK Sheet '{state}': {len(rows)} data rows, validations applied")

    # Write output file
    with open(OUTPUT_FILE, "wb") as f:
        f.write(output.getvalue())
    print(f"\nDone! File saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    asyncio.run(build())
