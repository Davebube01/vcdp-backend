"""
1. Inspect the NPMU tab in the AWPB file to understand its structure.
2. Copy/map NPMU data into Test_Bulk_Upload_Populated.xlsx.
3. Rebuild VCDP_Populated_Template.xlsx with the updated data.
"""
import asyncio
import io
import json
import os
import sys
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import selectinload

sys.path.append(os.path.abspath('.'))

AWPB_FILE = r"../excel/08_04_2025_VCDP_AWPB_2025_AF_1_2_NVCSC_APPROVED_ASAP _FINAL_v2.xlsx"
POPULATED_FILE = r"../excel/Test_Bulk_Upload_Populated.xlsx"
OUTPUT_FILE = r"../excel/VCDP_Populated_Template.xlsx"

TEMPLATE_COLUMNS = [
    "Ref ID", "Project / Activity Name", "Activity Name", "Activity Code", "Category / Costcode",
    "Record Type", "Status", "Institution Code", "Executing Agency", "FY Awarded",
    "FY Completed", "Programme Phase", "Fiscal Quarter", "LGA(s)", "Commodity",
    "VCDP Component", "VCDP Sub-Component(s)", "3FS Primary Component", "3FS Sub-Component(s)",
    "COFOG Code", "COFOG Division(s)", "COFOG Group(s)", "Funding Sources", "Sub Funding Sources",
    "Currency", "Exchange Rate", "Expenditure - FGN", "Expenditure - State",
    "Expenditure - IFAD Loan", "Expenditure - IFAD Grant", "Expenditure - OOF",
    "Expenditure - Beneficiary", "Expenditure - Private Sector",
    "Expenditure - Value Chain", "Expenditure - Other", "Expenditure - Total Reported",
    "Beneficiary Unit", "Beneficiaries - Total", "Beneficiaries - Male", "Beneficiaries - Female",
    "Beneficiaries - Youth (<35)", "Beneficiaries - PLWD", "Beneficiary Categories",
    "Quantity Q1", "Quantity Q2", "Quantity Q3", "Quantity Q4",
    "Value Chain Segment(s)", "Value Chain Segments (Other)", "Climate Aligned?",
    "Data Source(s)", "Classification Notes"
]

# Match the EXACT column names used in Test_Bulk_Upload_Populated.xlsx
# (may use em-dash or en-dash for FGN)
POPULATED_COLUMNS = [
    "Ref ID", "Project / Activity Name", "Activity Name", "Activity Code", "Category / Costcode",
    "Record Type", "Status", "Institution Code", "Executing Agency", "FY Awarded",
    "FY Completed", "Programme Phase", "Fiscal Quarter", "LGA(s)", "Commodity",
    "VCDP Component", "VCDP Sub-Component(s)", "3FS Primary Component", "3FS Sub-Component(s)",
    "COFOG Code", "COFOG Division(s)", "COFOG Group(s)", "Funding Sources", "Sub Funding Sources",
    "Currency", "Exchange Rate", "Expenditure \u2013 FGN", "Expenditure - State",
    "Expenditure - IFAD Loan", "Expenditure - IFAD Grant", "Expenditure - OOF",
    "Expenditure - Beneficiary", "Expenditure - Private Sector",
    "Expenditure - Value Chain", "Expenditure - Other", "Expenditure - Total Reported",
    "Beneficiary Unit", "Beneficiaries \u2013 Total", "Beneficiaries - Male", "Beneficiaries - Female",
    "Beneficiaries - Youth (<35)", "Beneficiaries - PLWD", "Beneficiary Categories",
    "Quantity Q1", "Quantity Q2", "Quantity Q3", "Quantity Q4",
    "Value Chain Segment(s)", "Value Chain Segments (Other)", "Climate Aligned?",
    "Data Source(s)", "Classification Notes"
]

VALID_STATE_TABS = [
    'Anambra', 'Benue', 'Ebonyi', 'Enugu', 'Kogi',
    'Nasarawa', 'Niger', 'Ogun', 'Taraba', 'NPMU'
]

VCDP_COMPONENTS = {
    "Component 1: Agricultural Market Development": [
        "Value Addition & Market Linkages", "Market Infrastructure",
        "Mainstreaming activities such as gender/youth, financial inclusion, E&CC and nutrition are cross-cutting",
    ],
    "Component 2: Smallholder Productivity Enhancement": [
        "Strengthening Farmers' Organisations", "Smallholder Production",
        "Mainstreaming activities such as gender/youth, financial inclusion, E&CC and nutrition are cross-cutting"
    ],
    "Component 3: Programme Management and Coordination": [
        "Procurement", "Knowledge Management & Sharing", "M&E", "Finance/Audit", "Admin",
    ],
}

THREEFS_COMPONENTS = {
    "1. Food Production": ["Production support (On-farm)", "Input supply & technologies", "Extension services"],
    "2. Food processing": ["Agro-processing", "Processing centers"],
    "3. Food social protection": ["Cash transfer schemes", "Emergency food assistance", "School feeding programs"],
    "4. Enabling environment": ["Rural roads or other transportation networks", "Market linkages"],
    "5. Governance": ["Climate adaptation", "Land management"],
}

FUNDING_SOURCES = {
    "Domestic Public Financing": ["FGN counterpart funding", "State/LGA contribution", "Beneficiary Contribution"],
    "International Development Financing": [
        "IFAD loan (Official Development Assistance (ODA))",
        "IFAD Grants (Official Development Assistance (ODA))", "Other Official Flows (OOF)"
    ],
    "Private Sector Financing": [],
    "Value Chain Financing": [],
}

VALUE_CHAIN_SEGMENTS = ["Production", "Processing", "Marketing", "Others"]
COMMODITIES = ["Rice", "Cassava", "Cross-cutting"]
FISCAL_YEARS = list(range(2013, 2051))


def fmt_list(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.startswith('['):
        try:
            items = json.loads(s)
            return ", ".join(str(i) for i in items)
        except:
            pass
    return s


def apply_dropdowns(worksheet, TEMPLATE_COLUMNS_LIST, ref_map, lga_ref_map, state):
    def add_val(col_name, formula):
        if not formula:
            return
        try:
            idx = TEMPLATE_COLUMNS_LIST.index(col_name) + 1
            col_letter = get_column_letter(idx)
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid Selection",
                error="Please select a value from the dropdown list."
            )
            dv.add(f"{col_letter}2:{col_letter}2000")
            worksheet.add_data_validation(dv)
        except ValueError:
            pass

    add_val("Record Type", ref_map.get("RecordType", ""))
    add_val("Status", ref_map.get("Status", ""))
    add_val("Currency", ref_map.get("Currency", ""))
    add_val("Climate Aligned?", ref_map.get("Climate", ""))
    add_val("Commodity", ref_map.get("Commodity", ""))
    add_val("Fiscal Quarter", ref_map.get("Quarters", ""))
    add_val("VCDP Component", ref_map.get("VCDP_Comp", ""))
    add_val("3FS Primary Component", ref_map.get("3FS_Comp", ""))
    add_val("FY Awarded", ref_map.get("Years", ""))
    add_val("FY Completed", ref_map.get("Years", ""))
    add_val("Beneficiary Unit", ref_map.get("Units", ""))
    add_val("Programme Phase", ref_map.get("Phases", ""))
    add_val("Funding Sources", ref_map.get("Funding", ""))
    add_val("Value Chain Segment(s)", ref_map.get("Segments", ""))
    add_val("LGA(s)", lga_ref_map.get(state, ""))


async def main():
    from app.database import AsyncSessionLocal
    from app.models import State

    # ── STEP 1: Inspect AWPB NPMU tab ─────────────────────────────────────────
    print("Reading AWPB file...")
    awpb = pd.ExcelFile(AWPB_FILE)
    print(f"  AWPB sheets: {awpb.sheet_names}")

    # Find the NPMU sheet (might be named differently)
    npmu_sheet = None
    for s in awpb.sheet_names:
        if "NPMU" in s.upper() or "FCT" in s.upper() or "NATIONAL" in s.upper():
            npmu_sheet = s
            break

    if not npmu_sheet:
        print("ERROR: Could not find NPMU/FCT sheet in AWPB file!")
        print(f"  Available sheets: {awpb.sheet_names}")
        return

    print(f"  Found NPMU sheet: '{npmu_sheet}'")
    df_awpb_npmu = pd.read_excel(awpb, sheet_name=npmu_sheet)
    print(f"  AWPB NPMU columns: {list(df_awpb_npmu.columns)}")
    print(f"  AWPB NPMU rows: {len(df_awpb_npmu)}")
    print(f"  First 3 rows sample:")
    print(df_awpb_npmu.head(3).to_string())

    # ── STEP 2: Map AWPB NPMU to populated template format ────────────────────
    print("\nMapping AWPB NPMU data to template columns...")
    # AWPB column mapping:
    #   "NATIONAL PROGRAMME..." (col A) = Project / Activity Name
    #   "Activity type code"            = Activity Code
    #   "Activity type name"            = Activity Name
    #   "Category / Costcode"           = Category / Costcode
    #   "Unit"                          = Beneficiary Unit
    #   "Quantity Q1/Q2/Q3/Q4"         = Quantity Q1-Q4
    #   "Institution"                   = Institution Code
    #   "Institution name"              = Executing Agency
    #   "Where"                         = LGA(s)
    #   "Budget Q1-Q4"                  = expenditure (quarterly budget)
    #   "Total all quarters"            = Expenditure - Total Reported
    #   "IFAD Loan (NAIRA)"             = Expenditure - IFAD Loan
    #   "IFAD Grant"                    = Expenditure - IFAD Grant
    #   "Federal Government of Nigeria" = Expenditure – FGN
    #   "State Government/LGAs"        = Expenditure - State
    #   "Beneficiaries"                 = Beneficiaries – Total
    #   "Budget comments"               = Classification Notes
    #   "Q1/Q2/Q3/Q4" (budget cols)    = quarterly budget breakdowns
    #   "Cat"                           = may indicate Component

    # Identify the main name column (first column)
    name_col = df_awpb_npmu.columns[0]

    npmu_rows = []
    for _, row in df_awpb_npmu.iterrows():
        def g(col, default=""):
            if col not in df_awpb_npmu.columns:
                return default
            v = row[col]
            return default if pd.isna(v) or str(v).strip() == "" else str(v).strip()

        project_name = g(name_col)
        activity_code = g("Activity type code")

        # Skip header/section rows (no activity code = section header)
        if not activity_code and not project_name:
            continue
        # Skip pure section/sub-section rows (no financial data at all)
        total = g("Total all quarters")
        if not activity_code and not total:
            continue

        new_row = {col: "" for col in POPULATED_COLUMNS}
        new_row["Project / Activity Name"] = project_name
        new_row["Activity Code"] = activity_code
        new_row["Activity Name"] = g("Activity type name")
        new_row["Category / Costcode"] = g("Category / Costcode")
        new_row["Beneficiary Unit"] = g("Unit", "Person")
        new_row["Institution Code"] = g("Institution")
        new_row["Executing Agency"] = g("Institution name")
        new_row["LGA(s)"] = g("Where")
        new_row["Record Type"] = "Budget"
        new_row["Status"] = "DRAFT"
        new_row["FY Awarded"] = "2025"
        new_row["FY Completed"] = "2025"
        new_row["Programme Phase"] = "2nd AF"
        new_row["Fiscal Quarter"] = "Q1, Q2, Q3, Q4"
        new_row["Currency"] = "NGN"
        new_row["Exchange Rate"] = "1"

        # Expenditure fields
        new_row["Expenditure \u2013 FGN"] = g("Federal Government of Nigeria", "0")
        new_row["Expenditure - State"] = g("State Government/LGAs", "0")
        new_row["Expenditure - IFAD Loan"] = g("IFAD Loan (NAIRA)", "0")
        new_row["Expenditure - IFAD Grant"] = g("IFAD Grant", "0")
        new_row["Expenditure - Beneficiary"] = "0"
        new_row["Expenditure - OOF"] = "0"
        new_row["Expenditure - Private Sector"] = "0"
        new_row["Expenditure - Value Chain"] = "0"
        new_row["Expenditure - Other"] = "0"
        new_row["Expenditure - Total Reported"] = g("Total all financiers") or g("Total all quarters", "0")

        # Quantities
        new_row["Quantity Q1"] = g("Quantity Q1", "0")
        new_row["Quantity Q2"] = g("Quantity Q2", "0")
        new_row["Quantity Q3"] = g("Quantity Q3", "0")
        new_row["Quantity Q4"] = g("Quantity Q4", "0")

        # Beneficiaries
        new_row["Beneficiaries \u2013 Total"] = g("Beneficiaries", "0")

        # Notes
        new_row["Classification Notes"] = g("Budget comments")

        npmu_rows.append(new_row)

    print(f"  Mapped {len(npmu_rows)} NPMU rows")

    # ── STEP 3: Add NPMU sheet to Test_Bulk_Upload_Populated.xlsx ─────────────
    print("\nUpdating Test_Bulk_Upload_Populated.xlsx with NPMU sheet...")
    populated = pd.ExcelFile(POPULATED_FILE)
    existing_sheets = populated.sheet_names
    print(f"  Existing sheets: {existing_sheets}")

    output_populated = io.BytesIO()
    with pd.ExcelWriter(output_populated, engine='openpyxl') as writer:
        # Copy all existing sheets
        for sheet in existing_sheets:
            df_existing = pd.read_excel(populated, sheet_name=sheet)
            df_existing.to_excel(writer, index=False, sheet_name=sheet)

        # Add NPMU sheet
        df_npmu = pd.DataFrame(npmu_rows, columns=POPULATED_COLUMNS)
        df_npmu.to_excel(writer, index=False, sheet_name="NPMU")

    with open(POPULATED_FILE, "wb") as f:
        f.write(output_populated.getvalue())
    print(f"  Saved updated populated file with NPMU tab ({len(npmu_rows)} rows)")

    # ── STEP 4: Rebuild VCDP_Populated_Template.xlsx ──────────────────────────
    print("\nRebuilding VCDP_Populated_Template.xlsx with all state data...")

    # Reload all data
    src = pd.ExcelFile(POPULATED_FILE)
    data_by_state = {}
    for sheet in src.sheet_names:
        if sheet not in VALID_STATE_TABS:
            continue
        df = pd.read_excel(src, sheet_name=sheet)
        if df.empty:
            continue
        data_by_state[sheet] = df.to_dict('records')
        print(f"  Loaded {len(data_by_state[sheet])} rows from '{sheet}'")

    # Get LGAs from DB
    async with AsyncSessionLocal() as db:
        result = await db.execute(select(State).options(selectinload(State.lgas)))
        db_states = result.scalars().all()
        state_lga_map = {s.name: [l.name for l in s.lgas] for s in db_states}
        if "FCT (NPMU)" in state_lga_map:
            state_lga_map["NPMU"] = state_lga_map["FCT (NPMU)"]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        options_ws = workbook.create_sheet("Options")
        options_ws.sheet_state = 'hidden'

        options_map = {
            "RecordType": ["Actual", "Budget"],
            "Status": ["DRAFT", "PENDING", "PUBLISHED", "REJECTED"],
            "Currency": ["NGN", "USD"],
            "Climate": ["Yes", "No"],
            "Commodity": COMMODITIES,
            "Quarters": ["Q1", "Q2", "Q3", "Q4", "Q1, Q2, Q3, Q4"],
            "VCDP_Comp": list(VCDP_COMPONENTS.keys()),
            "3FS_Comp": list(THREEFS_COMPONENTS.keys()),
            "Years": [str(y) for y in FISCAL_YEARS],
            "Units": ["Person", "Hectares", "Kilometers", "Tons", "Number", "Other"],
            "Phases": ["Original (2013-2018)", "1st AF", "2nd AF"],
            "Funding": list(FUNDING_SOURCES.keys()),
            "Segments": VALUE_CHAIN_SEGMENTS,
        }

        col_idx = 1
        ref_map = {}
        for key, vals in options_map.items():
            for row_idx, val in enumerate(vals, 1):
                options_ws.cell(row=row_idx, column=col_idx, value=val)
            ref_map[key] = f"'Options'!${get_column_letter(col_idx)}$1:${get_column_letter(col_idx)}${len(vals)}"
            col_idx += 1

        lga_ref_map = {}
        for state_name, lgas in state_lga_map.items():
            if not lgas:
                continue
            for row_idx, lga in enumerate(lgas, 1):
                options_ws.cell(row=row_idx, column=col_idx, value=lga)
            lga_ref_map[state_name] = f"'Options'!${get_column_letter(col_idx)}$1:${get_column_letter(col_idx)}${len(lgas)}"
            col_idx += 1

        for state in VALID_STATE_TABS:
            rows = data_by_state.get(state, [])
            if rows:
                df_out = pd.DataFrame(rows)
                # Ensure all template columns exist
                for col in POPULATED_COLUMNS:
                    if col not in df_out.columns:
                        df_out[col] = ""
                df_out = df_out.reindex(columns=POPULATED_COLUMNS, fill_value="")
            else:
                df_out = pd.DataFrame(columns=POPULATED_COLUMNS)

            df_out.to_excel(writer, index=False, sheet_name=state)
            worksheet = writer.sheets[state]

            apply_dropdowns(worksheet, POPULATED_COLUMNS, ref_map, lga_ref_map, state)

            worksheet.freeze_panes = "A2"
            for col in worksheet.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value),
                    default=8
                )
                worksheet.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

            print(f"  Sheet '{state}': {len(rows)} rows")

    with open(OUTPUT_FILE, "wb") as f:
        f.write(output.getvalue())
    print(f"\nDone! Saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    asyncio.run(main())
