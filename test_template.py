import asyncio
import io
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime
from sqlalchemy import select
from sqlalchemy.orm import selectinload

# Mocking parts of the app
TEMPLATE_COLUMNS = [
    "Ref ID", "Project / Activity Name", "Activity Name", "Activity Code", "Category / Costcode",
    "Record Type", "Status", "Institution Code", "Executing Agency", "FY Awarded",
    "FY Completed", "Programme Phase", "Fiscal Quarter", "LGA(s)", "Commodity",
    "VCDP Component", "VCDP Sub-Component(s)", "3FS Primary Component", "3FS Sub-Component(s)",
    "COFOG Code", "COFOG Division(s)", "COFOG Group(s)", "Funding Sources", "Sub Funding Sources",
    "Currency", "Exchange Rate", "Expenditure – FGN", "Expenditure - State", "Expenditure - IFAD Loan",
    "Expenditure - IFAD Grant", "Expenditure - OOF", "Expenditure - Beneficiary", "Expenditure - Private Sector",
    "Expenditure - Value Chain", "Expenditure - Other", "Expenditure - Total Reported", "Beneficiary Unit",
    "Beneficiaries – Total", "Beneficiaries - Male", "Beneficiaries - Female", "Beneficiaries - Youth (<35)",
    "Beneficiaries - PLWD", "Beneficiary Categories", "Quantity Q1", "Quantity Q2", "Quantity Q3", "Quantity Q4",
    "Value Chain Segment(s)", "Value Chain Segments (Other)", "Climate Aligned?", "Data Source(s)", "Classification Notes"
]

VALID_STATE_TABS = ['Anambra', 'NPMU']

async def test_template_gen():
    from app.database import AsyncSessionLocal
    from app.models import State, User, UserRole
    from app.routers.meta import (
        VCDP_COMPONENTS, THREEFS_COMPONENTS, COMMODITIES, FISCAL_YEARS
    )

    async with AsyncSessionLocal() as db:
        # Mock user
        current_user = User(email="test@vcdp.org", role=UserRole.NATIONAL_ADMIN)
        
        output = io.BytesIO()
        result = await db.execute(select(State).options(selectinload(State.lgas)))
        db_states = result.scalars().all()
        state_lga_map = {s.name: [l.name for l in s.lgas] for s in db_states}
        if "FCT (NPMU)" in state_lga_map:
            state_lga_map["NPMU"] = state_lga_map["FCT (NPMU)"]

        df = pd.DataFrame(columns=TEMPLATE_COLUMNS)
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            workbook = writer.book
            options_ws = workbook.create_sheet("Options")
            options_ws.sheet_state = 'hidden'
            
            options_map = {
                "RecordType": ["Actual", "Budget"],
                "Status": ["DRAFT", "PENDING", "PUBLISHED", "REJECTED"],
                "Currency": ["NGN", "USD"],
                "Climate": ["Yes", "No"],
                "Commodity": COMMODITIES,
                "Quarters": ["Q1", "Q2", "Q3", "Q4", "Q1, Q2, Q3, Q4"],
                "VCDP_Comp": list(VCDP_COMPONENTS.keys()),
                "3FS_Comp": list(THREEFS_COMPONENTS.keys()),
                "Years": [str(y) for y in FISCAL_YEARS],
                "Units": ["Person", "Hectares", "Kilometers", "Tons", "Number", "Other"],
                "Phases": ["Original (2013-2018)", "1st AF", "2nd AF"]
            }
            
            col_idx = 1
            ref_map = {}
            for key, vals in options_map.items():
                for row_idx, val in enumerate(vals, 1):
                    options_ws.cell(row=row_idx, column=col_idx, value=val)
                ref_map[key] = f"'Options'!${get_column_letter(col_idx)}$1:${get_column_letter(col_idx)}${len(vals)}"
                col_idx += 1
            
            lga_ref_map = {}
            for state_name, lgas in state_lga_map.items():
                if not lgas: continue
                for row_idx, lga in enumerate(lgas, 1):
                    options_ws.cell(row=row_idx, column=col_idx, value=lga)
                lga_ref_map[state_name] = f"'Options'!${get_column_letter(col_idx)}$1:${get_column_letter(col_idx)}${len(lgas)}"
                col_idx += 1

            for state in VALID_STATE_TABS:
                df.to_excel(writer, index=False, sheet_name=state)
                worksheet = writer.sheets[state]
                
                def add_val(col_name, formula):
                    try:
                        idx = TEMPLATE_COLUMNS.index(col_name) + 1
                        col_letter = get_column_letter(idx)
                        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                        dv.add(f"{col_letter}2:{col_letter}1000")
                        worksheet.add_data_validation(dv)
                    except Exception as e:
                        print(f"Error {col_name}: {e}")

                add_val("Record Type", ref_map["RecordType"])
                add_val("LGA(s)", lga_ref_map.get(state, ""))

        with open("test_template.xlsx", "wb") as f:
            f.write(output.getvalue())
        print("Template generated as test_template.xlsx")

if __name__ == "__main__":
    import os
    import sys
    sys.path.append(os.path.abspath('.'))
    asyncio.run(test_template_gen())
